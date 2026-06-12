#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
读取综维数据.xlsx，输出为智能体可读的结构化文本（极简表格版）
"""

import openpyxl

INPUT_FILE = "综维数据.xlsx"
OUTPUT_FILE = "综维数据分析.txt"


def fmt(val):
    """格式化数值"""
    if isinstance(val, (int, float)):
        return f"{val:.2f}"
    return str(val)


def main():
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)

    sheet_names = wb.sheetnames
    # 映射sheet名：从实际sheet名到友好名称
    name_map = {}
    for sn in sheet_names:
        if "PON" in sn.upper() or "pon" in sn.lower() or "扣" in sn:
            name_map[sn] = "PON口不可用"
        elif "故障处理" in sn or "时长" in sn:
            name_map[sn] = "平均故障处理时长"
        elif "故障" in sn or "工单" in sn:
            name_map[sn] = "人均故障工单"
        else:
            name_map[sn] = sn

    lines = []

    for sheet_idx, sheet_name in enumerate(sheet_names):
        ws = wb[sheet_name]
        display_name = name_map.get(sheet_name, sheet_name)

        # 读取数据
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header = list(rows[0])
        months = [str(h).strip() for h in header[1:]]

        # 收集所有区县数据
        data = []
        for row in rows[1:]:
            if row[0] is None:
                continue
            region = str(row[0]).strip()
            values = []
            for v in row[1:]:
                try:
                    values.append(float(v) if v is not None else 0.0)
                except (ValueError, TypeError):
                    values.append(0.0)
            data.append((region, values))

        if not data:
            continue

        # 指标标题行
        time_range = f"{months[0]} ~ {months[-1]}" if months else "未知"
        lines.append(f"{display_name} | {time_range}")
        lines.append("")

        # 表格头部
        cols = ["区县"] + months
        lines.append("| " + " | ".join(cols) + " |")
        lines.append("| " + " | ".join(["---"] * len(cols)) + " |")

        # 数据行
        for region, values in data:
            row_str = " | ".join([region] + [fmt(v) for v in values])
            lines.append("| " + row_str + " |")

        lines.append("")

    # 写入文件
    output = "\n".join(lines)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(output)

    print(f"✅ 已生成：{OUTPUT_FILE}（共 {len(lines)} 行）")
    print("\n" + "=" * 60)
    print("预览：")
    print("=" * 60)
    preview = "\n".join(lines[:40])
    print(preview)


if __name__ == "__main__":
    main()